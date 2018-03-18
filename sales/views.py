import json
import locale
from datetime import datetime, date, timedelta

from django.contrib.auth.decorators import login_required, permission_required
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment

from cloudkitchen.settings.base import PAGE_TITLE
from products.models import Cartridge, PackageCartridge, PackageCartridgeRecipe
from sales.models import TicketBase, TicketPOS, CartridgeTicketDetail, PackageCartridgeTicketDetail
from users.models.users import User as UserProfile
from helpers.sales_helper import TicketPOSHelper
from helpers.products_helper import ProductsHelper
from helpers.helpers import Helper


# -------------------------------------  Sales -------------------------------------
class SalesReport(TemplateView):
    def get(self, request, *args, **kwargs):
        helpers = Helper()
        sales_helper = TicketPOSHelper()
        products_helper = ProductsHelper()
        workbook = Workbook()
        count = 4

        ws1 = workbook.active
        ws1.title = "Reporte de Ventas"
        ws1['B1'] = 'Reporte General de Ventas'
        ws1['B1'].alignment = Alignment(horizontal='center')
        ws1.merge_cells('B1:R1')
        ws1['A3'] = 'ID'
        ws1['B3'] = 'ID Venta'
        ws1['C3'] = 'Núm. Orden'
        ws1['D3'] = 'Fecha'
        ws1['E3'] = 'Hora'
        ws1['F3'] = 'Horas'
        ws1['G3'] = 'ID Paquete'
        ws1['H3'] = 'Paquete'
        ws1['I3'] = 'Roti'
        ws1['J3'] = 'Ensalada'
        ws1['K3'] = 'Licuados'
        ws1['L3'] = 'Fruta'
        ws1['M3'] = 'Jugo'
        ws1['N3'] = 'Agua'
        ws1['O3'] = 'Cantidad'
        ws1['P3'] = 'Tipo de Pago'
        ws1['Q3'] = 'Precio Unitario'
        ws1['R3'] = 'Total'

        for cartridge_ticket_detail in sales_helper.get_cartridges_tickets_details():
            created_at = helpers.naive_to_datetime(cartridge_ticket_detail.ticket_base.created_at)
            cartridge_name = cartridge_ticket_detail.cartridge.name
            subcategory = cartridge_ticket_detail.cartridge.subcategory
            payment_type = cartridge_ticket_detail.ticket_base.payment_type

            ws1.cell(row=count, column=1, value=cartridge_ticket_detail.id)
            ws1.cell(row=count, column=2, value=cartridge_ticket_detail.ticket_base.id)
            ws1.cell(row=count, column=3, value=cartridge_ticket_detail.ticket_base.order_number)
            ws1.cell(row=count, column=4, value=created_at.date())
            ws1.cell(row=count, column=5, value=created_at.time())
            ws1.cell(row=count, column=5).number_format = 'hh:mm:ss AM/PM'
            ws1.cell(row=count, column=6, value=created_at.time())
            ws1.cell(row=count, column=6).number_format = 'hh AM/PM'
            ws1.cell(row=count, column=9, value=cartridge_name if subcategory == 'RO' else '')
            ws1.cell(row=count, column=10, value=cartridge_name if subcategory == 'SA' else '')
            ws1.cell(row=count, column=11, value=cartridge_name if subcategory == 'SM' else '')
            ws1.cell(row=count, column=12, value=cartridge_name if subcategory == 'FR' else '')
            ws1.cell(row=count, column=13, value=cartridge_name if subcategory == 'JU' else '')
            ws1.cell(row=count, column=14, value=cartridge_name if subcategory == 'WA' else '')
            ws1.cell(row=count, column=15, value=cartridge_ticket_detail.quantity)
            ws1.cell(row=count, column=16, value='Efectivo' if payment_type == 'CA' else 'Tarjeta')
            ws1.cell(row=count, column=17, value=cartridge_ticket_detail.cartridge.price)
            ws1.cell(row=count, column=18, value=cartridge_ticket_detail.cartridge.price * cartridge_ticket_detail.quantity)

            count += 1

        # Only Packages

        for package_ticket_detail in sales_helper.get_packages_tickets_details():
            created_at = helpers.naive_to_datetime(package_ticket_detail.ticket_base.created_at)

            payment_type = package_ticket_detail.ticket_base.payment_type
            ws1.cell(row=count, column=1, value=package_ticket_detail.id)
            ws1.cell(row=count, column=2, value=package_ticket_detail.ticket_base.id)
            ws1.cell(row=count, column=3, value=package_ticket_detail.ticket_base.order_number)
            ws1.cell(row=count, column=4, value=created_at.date())
            ws1.cell(row=count, column=5, value=created_at.time())
            ws1.cell(row=count, column=5).number_format = 'hh:mm:ss AM/PM'
            ws1.cell(row=count, column=6, value=created_at.time())
            ws1.cell(row=count, column=6).number_format = 'HH AM/PM'
            ws1.cell(row=count, column=7, value=package_ticket_detail.package_cartridge.id)
            ws1.cell(row=count, column=8, value=package_ticket_detail.package_cartridge.name)

            # Fill cartridge rows
            packages = products_helper.get_packages_cartridges_recipes().filter(
                package_cartridge=package_ticket_detail.package_cartridge)
                
            for package in packages:
                subcategory = package.cartridge.subcategory
                cartridge_name = package.cartridge.name
                if subcategory == 'RO':
                    ws1.cell(row=count, column=9, value=cartridge_name)
                if subcategory == 'FR':
                    ws1.cell(row=count, column=12, value=cartridge_name)
                if subcategory == 'JU':
                    ws1.cell(row=count, column=13, value=cartridge_name)
                if subcategory == 'WA':
                    ws1.cell(row=count, column=14, value=cartridge_name)

            ws1.cell(row=count, column=15, value=package_ticket_detail.quantity)
            ws1.cell(row=count, column=16, value='Efectivo' if payment_type == 'CA' else 'Tarjeta')
            ws1.cell(row=count, column=17, value=package_ticket_detail.package_cartridge.price)
            ws1.cell(row=count, column=18, value=package_ticket_detail.price * package_ticket_detail.quantity)

            count += 1
        
        locale.setlocale(locale.LC_ALL, '')
        file_name = 'ReporteGralVentas-{0}.xlsx'.format(datetime.now().strftime("%Y%m%d%H%M%S"))
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename="{0}"'.format(file_name)
        response['content-disposition'] = content
        workbook.save(response)
        return response


@permission_required('users.can_see_sales')
def sales(request):
    locale.setlocale(locale.LC_ALL, '')
    sales_helper = TicketPOSHelper()
    helper = Helper()

    if request.method == 'POST':
        if request.POST['type'] == 'dates_range':
            dates_range = sales_helper.get_dates_range()
            return JsonResponse(dates_range)
        if request.POST['type'] == 'sales_day':
            """
            Returns a list with objects:
            Each object has the following characteristics
            """
            sales_day_list = []
            start_day = helper.naive_to_datetime(datetime.strptime(request.POST['date'], '%d-%m-%Y').date())
            end_date = helper.naive_to_datetime(start_day + timedelta(days=1))
            
            tickets_objects = sales_helper.tickets_pos.filter(ticket__created_at__range=[start_day, end_date])

            for ticket_pos in tickets_objects:
                """
                Filling in the sales list of the day
                """
                earnings_sale_object = {
                    'id_ticket': ticket_pos.ticket.id,
                    'datetime': timezone.localtime(ticket_pos.ticket.created_at),
                    'earnings': 0
                }

                # Cartridge Ticket Detail
                for cartridge_ticket_detail in sales_helper.get_cartridges_tickets_details():
                    if cartridge_ticket_detail.ticket_base == ticket_pos.ticket:
                        earnings_sale_object['earnings'] += cartridge_ticket_detail.price

                # Package Ticket Detail
                for package_ticket_detail in sales_helper.get_packages_tickets_details():
                    if package_ticket_detail.ticket_base == ticket_pos.ticket:
                        earnings_sale_object['earnings'] += package_ticket_detail.price

                sales_day_list.append(earnings_sale_object)
            return JsonResponse({'sales_day_list': sales_day_list})

        if request.POST['type'] == 'ticket_details':
            ticket_id = int(request.POST['ticket_id'])
            ticket_object = {
                'ticket_id': ticket_id,
                'ticket_order': '',
                'cartridges': [],
                'packages': [],
            }

            # Cartridge Ticket Details
            for cartridge_ticket_detail in sales_helper.get_cartridges_tickets_details():
                if cartridge_ticket_detail.ticket_base.id == ticket_id:
                    ticket_object['ticket_order'] = cartridge_ticket_detail.ticket_base.order_number

                    cartridge_object = {
                        'name': cartridge_ticket_detail.cartridge.name,
                        'quantity': cartridge_ticket_detail.quantity,
                        'total': cartridge_ticket_detail.price
                    }

                    ticket_object['cartridges'].append(cartridge_object)

            # Package Ticket Details
            for package_ticket_detail in sales_helper.get_packages_tickets_details():
                if package_ticket_detail.ticket_base.id == ticket_id:
                    ticket_object['ticket_order'] = package_ticket_detail.ticket_base.order_number
                    cartridges_list = []

                    package_cartridge_recipe = PackageCartridgeRecipe.objects.filter(
                        package_cartridge=package_ticket_detail.package_cartridge)

                    for cartridge_recipe in package_cartridge_recipe:
                        cartridges_list.append(cartridge_recipe.cartridge.name)

                    package_cartridge_object = {
                        'cartridges': cartridges_list,
                        'quantity': package_ticket_detail.quantity,
                        'total': package_ticket_detail.price
                    }

                    ticket_object['packages'].append(package_cartridge_object)

            return JsonResponse({'ticket_details': ticket_object})

        if request.POST['type'] == 'tickets':
            tickets_objects_list = []

            for ticket_pos in sales_helper.tickets_pos:
                for cartridge_ticket_detail in sales_helper.get_cartridges_tickets_details():
                    if cartridge_ticket_detail.ticket == ticket_pos.ticket:
                        ticket_object = {
                            'ID': ticket_pos.ticket.id,
                            'Fecha': timezone.localtime(ticket_pos.ticket.created_at).date(),
                            'Hora': timezone.localtime(ticket_pos.ticket.created_at).time(),
                            'Vendedor': ticket_pos.cashier.username,
                        }
                        if ticket_pos.ticket.payment_type == 'CA':
                            ticket_object['Tipo de Pago'] = 'Efectivo'
                        else:
                            ticket_object['Tipo de Pago'] = 'Crédito'
                        if cartridge_ticket_detail.cartridge:
                            ticket_object['Producto'] = cartridge_ticket_detail.cartridge.name
                        else:
                            ticket_object['Producto'] = None
                        if cartridge_ticket_detail.package_cartridge:
                            ticket_object['Paquete'] = cartridge_ticket_detail.package_cartridge.name
                        else:
                            ticket_object['Paquete'] = None
                        ticket_object['Cantidad'] = cartridge_ticket_detail.quantity
                        ticket_object['Total'] = cartridge_ticket_detail.price
                        ticket_object[
                            'Precio Unitario'] = cartridge_ticket_detail.price / cartridge_ticket_detail.quantity

                        tickets_objects_list.append(ticket_object)

            return JsonResponse({'ticket': tickets_objects_list})

        if request.POST['type'] == 'sales_week':
            initial_date = request.POST['dt_week'].split(',')[0]
            final_date = request.POST['dt_week'].split(',')[1]
            initial_date = helper.parse_to_datetime(initial_date)
            final_date = helper.parse_to_datetime(final_date) + timedelta(days=1)
            week_sales = sales_helper.get_sales_list(initial_date, final_date)
            tickets = sales_helper.get_tickets_list(initial_date, final_date)
            data = {
                'sales': week_sales,
                'tickets': tickets,
                'week_number': helper.get_week_number(initial_date)
            }
            return JsonResponse(data)

        if request.POST['type'] == 'sales_actual_week':
            sales_actual_week = sales_helper.get_sales_actual_week()
            return JsonResponse(sales_actual_week)

    # Any other request method:
    template = 'sales/sales.html'
    title = 'Registro de Ventas'

    initial_date, final_date = helper.get_initial_final_week_datetime()

    context = {
        'title': PAGE_TITLE + ' | ' + title,
        'page_title': title,
        'actual_year': datetime.now().year,
        'today_name': datetime.now().strftime("%A"),
        'today_number': datetime.now().strftime("%w"),
        'week_number': helper.get_week_number(date.today()),
        'tickets': sales_helper.get_tickets_list(initial_date, final_date),
    }

    return render(request, template, context)


@login_required(login_url='users:login')
def delete_sale(request):
    if request.method == 'POST':
        ticket_id = request.POST['ticket_id']
        ticket = TicketBase.objects.get(id=ticket_id)
        ticket.delete()
        return JsonResponse({'result': 'success'})


@permission_required('users.can_sells')
def new_sale(request):
    template = 'sales/new_sale_v2.html'
    title = 'Nueva venta'
    context = {
        'page_title': PAGE_TITLE,
        'title': title,
    }

    return render(request, template, context)


@permission_required('users.can_sell')
def new_sale__(request):
    helper = Helper()
    sales_helper = TicketPOSHelper()
    products_helper = ProductsHelper()
    if request.method == 'POST':
        if request.POST['ticket']:
            username = request.user
            user_profile_object = get_object_or_404(UserProfile, username=username)
            ticket_detail_json_object = json.loads(request.POST.get('ticket'))
            payment_type = ticket_detail_json_object['payment_type']
            new_ticket_base = TicketBase(
                order_number=sales_helper.get_new_order_number(),
                payment_type=payment_type,
            )
            new_ticket_base.save()
            new_ticket_pos = TicketPOS(
                ticket=new_ticket_base,
                cashier=user_profile_object,
            )
            new_ticket_pos.save()

            """
            Saves the tickets details for cartridges
            """
            for ticket_detail in ticket_detail_json_object['cartridges']:
                cartridge = get_object_or_404(Cartridge, id=ticket_detail['id'])
                quantity = ticket_detail['quantity']
                price = ticket_detail['price']
                new_cartridge_ticket_detail = CartridgeTicketDetail(
                    ticket_base=new_ticket_base,
                    cartridge=cartridge,
                    quantity=quantity,
                    price=price
                )
                new_cartridge_ticket_detail.save()

            """
            Saves the tickets details for packages
            """
            for ticket_detail_packages in ticket_detail_json_object['packages']:
                quantity = ticket_detail_packages['quantity']
                price = ticket_detail_packages['price']
                packages_id_list = ticket_detail_packages['id_list']
                package_id = None
                is_new_package = True
                packages_recipes = PackageCartridge.objects.all()

                for package_recipe in packages_recipes:
                    """
                    Gets the cartridges for each package cartridge and compares
                    each package recipe cartridges if is equal that packages_id_list
                    """
                    cartridges_per_recipe = PackageCartridgeRecipe.objects.select_related(
                        'package_cartridge', 'cartridge').filter(package_cartridge=package_recipe)
                    cartridges_in_package_recipe = []

                    for cartridge_recipe in cartridges_per_recipe:
                        cartridges_in_package_recipe.append(cartridge_recipe.cartridge.id)

                    if helper.are_equal_lists(cartridges_in_package_recipe, packages_id_list):
                        is_new_package = False
                        package_id = package_recipe.id

                if is_new_package:
                    package_name = ticket_detail_packages['name']
                    package_price = ticket_detail_packages['price']
                    new_package_cartridge = PackageCartridge(name=package_name, price=package_price, is_active=True)
                    new_package_cartridge.save()

                    """
                    Creates a new package
                    """
                    for id_cartridge in packages_id_list:
                        cartridge = get_object_or_404(Cartridge, id=id_cartridge)
                        new_package_recipe_object = PackageCartridgeRecipe(
                            package_cartridge=new_package_cartridge,
                            cartridge=cartridge,
                            quantity=1
                        )
                        new_package_recipe_object.save()

                    """
                    Creates the ticket detail
                    """
                    new_cartridge_ticket_detail = PackageCartridgeTicketDetail(
                        ticket_base=new_ticket_base,
                        package_cartridge=new_package_cartridge,
                        quantity=quantity,
                        price=price
                    )
                    new_cartridge_ticket_detail.save()

                else:
                    """
                    Uses an existent package
                    """
                    package_cartridge = get_object_or_404(PackageCartridge, id=package_id)
                    new_cartridge_ticket_detail = PackageCartridgeTicketDetail(
                        ticket_base=new_ticket_base,
                        package_cartridge=package_cartridge,
                        quantity=quantity,
                        price=price,
                    )
                    new_cartridge_ticket_detail.save()

            json_response = {
                'status': 'ready',
                'ticket_id': new_ticket_base.id,
                'ticket_order': new_ticket_base.order_number,
            }
            return JsonResponse(json_response)

        return JsonResponse({'status': 'error'})

    else:
        cartridges_list = products_helper.cartridges
        package_cartridges = PackageCartridge.objects.all()
        template = 'sales/new_sale.html'
        title = 'Nueva venta'
        context = {
            'page_title': PAGE_TITLE,
            'title': title,
            'cartridges': cartridges_list,
            'package_cartridges': package_cartridges
        }
        return render(request, template, context)


# -------------------------------- Test ------------------------------
def test_sales_update(request):
    template = 'sales/test.html'

    return render(request, template)
